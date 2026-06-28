import os
import re
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date


# ── Color palette ────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"   # header backgrounds
TEAL        = "00B4D8"   # accent / section headers
LIGHT_GRAY  = "F2F4F7"   # alternating rows / fillable cells
WHITE       = "FFFFFF"
RED_WARN    = "C0392B"   # safety warning text
YELLOW_FILL = "FFF9C4"   # cells the technician fills in
GREEN_FILL  = "E8F5E9"   # verify checkpoint rows
ORANGE_FILL = "FFF3E0"   # point of no return rows


# ── Border helpers ────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="444444")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Cell writer helpers ───────────────────────────────────────────────────────
def header_cell(ws, row, col, text, span=1, font_size=12, bg=DARK_NAVY, fg=WHITE):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=fg, size=font_size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    cell.border = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def label_cell(ws, row, col, text, bold=True, bg=LIGHT_GRAY, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, size=10, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    cell.border = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def fill_cell(ws, row, col, text="", span=1, bg=YELLOW_FILL, font_size=10):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(size=font_size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    cell.border = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def warn_cell(ws, row, col, text, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=RED_WARN, size=10, name="Arial")
    cell.fill = PatternFill("solid", start_color="FDECEA")
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
    cell.border = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def section_header(ws, row, text, num_cols=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=WHITE, size=11, name="Arial")
    cell.fill = PatternFill("solid", start_color=TEAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thick_border()
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=num_cols)
    ws.row_dimensions[row].height = 22


# ── Parse markdown MOP into sections ─────────────────────────────────────────
def parse_mop_markdown(md_text):
    """
    Extracts key sections from the generated markdown MOP.
    Handles both plain ALL CAPS headers and markdown bold/heading formatted headers.
    Returns a dict keyed by section name.
    """
    sections = {}
    current_key = None
    current_lines = []

    # Known section keys to look for
    known_sections = [
        "STEP BY STEP PROCEDURE",
        "SAFETY WARNINGS",
        "PREREQUISITES",
        "REQUIRED PPE",
        "EXPECTED OUTCOME",
        "IF SOMETHING GOES WRONG",
        "ESCALATION CONTACTS",
        "APPENDIX A",
        "APPENDIX B",
        "APPENDIX C",
        "APPENDIX D",
        "APPENDIX E",
    ]

    for line in md_text.splitlines():
        stripped = line.strip()
        # Strip markdown formatting for header detection
        clean = re.sub(r'[#*_`]', '', stripped).strip()

        matched_section = None
        for section in known_sections:
            if clean.upper().startswith(section):
                matched_section = section
                break

        # Also match plain ALL CAPS lines
        if not matched_section and re.match(r'^[A-Z][A-Z\s/\-]+:?\s*$', clean) and len(clean) > 3:
            matched_section = clean.rstrip(":")

        if matched_section:
            if current_key:
                sections[current_key] = "\n".join(current_lines).strip()
            current_key = matched_section
            current_lines = []
        else:
            current_lines.append(line)

    if current_key:
        sections[current_key] = "\n".join(current_lines).strip()

    return sections


def extract_steps(steps_text):
    """
    Parses numbered steps from the STEP BY STEP PROCEDURE section.
    Handles formats:
      - Plain:  1. Action text
      - Bold:   **Step 1.** Action text
      - Bold:   Step 1. Action text
    Also captures VERIFY CHECKPOINTs and POINTs OF NO RETURN as special rows.
    Returns list of dicts with keys: number, text, is_verify, is_ponr, is_photo, is_phase
    """
    steps = []
    current_step = None

    for line in steps_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        # Remove markdown bold markers for matching
        clean = re.sub(r'\*+', '', stripped).strip()

        # Detect PHASE headers
        if re.match(r'^PHASE\s+\d+', clean, re.IGNORECASE):
            if current_step:
                steps.append(current_step)
                current_step = None
            steps.append({
                "number": None,
                "text": clean,
                "is_verify": False,
                "is_ponr": False,
                "is_photo": False,
                "is_phase": True,
            })
            continue

        # Detect VERIFY CHECKPOINT lines
        if re.match(r'^VERIFY CHECKPOINT', clean, re.IGNORECASE):
            if current_step:
                steps.append(current_step)
                current_step = None
            steps.append({
                "number": None,
                "text": clean,
                "is_verify": True,
                "is_ponr": False,
                "is_photo": False,
                "is_phase": False,
            })
            continue

        # Detect POINT OF NO RETURN lines
        if re.match(r'^POINT OF NO RETURN', clean, re.IGNORECASE):
            if current_step:
                steps.append(current_step)
                current_step = None
            steps.append({
                "number": None,
                "text": clean,
                "is_verify": False,
                "is_ponr": True,
                "is_photo": False,
                "is_phase": False,
            })
            continue

        # Detect photo placeholder lines
        if re.match(r'^\[PHOTO', clean, re.IGNORECASE):
            if current_step:
                steps.append(current_step)
                current_step = None
            steps.append({
                "number": None,
                "text": clean,
                "is_verify": False,
                "is_ponr": False,
                "is_photo": True,
                "is_phase": False,
            })
            continue

        # Match numbered steps in any format:
        # "1. text" or "Step 1. text" or "Step 1: text"
        match = re.match(r'^(?:Step\s+)?(\d+)[.:]\s+(.*)', clean, re.IGNORECASE)
        if match:
            if current_step:
                steps.append(current_step)
            text = match.group(2).strip()
            current_step = {
                "number": int(match.group(1)),
                "text": text,
                "is_verify": "VERIFY" in text.upper(),
                "is_ponr": "POINT OF NO RETURN" in text.upper(),
                "is_photo": "PHOTO REQUIRED" in text.upper() or "PHOTO:" in text.upper(),
                "is_phase": False,
            }
        elif current_step:
            # Continuation of current step
            current_step["text"] += " " + clean

    if current_step:
        steps.append(current_step)

    return steps


def extract_warnings(warnings_text):
    lines = [l.strip().lstrip("*").strip()
             for l in warnings_text.splitlines()
             if l.strip() and l.strip() not in ("**SAFETY WARNINGS:**", "SAFETY WARNINGS:")]
    return [l for l in lines if l]


# ── Sheet 1: Cover Page ───────────────────────────────────────────────────────
def build_cover_sheet(ws, meta):
    ws.title = "Cover Page"
    ws.sheet_view.showGridLines = False

    for col, width in zip(range(1, 9), [18, 22, 18, 18, 18, 18, 18, 18]):
        ws.column_dimensions[get_column_letter(col)].width = width

    r = 1
    # Title banner
    cell = ws.cell(row=r, column=1, value="METHOD OF PROCEDURE")
    cell.font = Font(bold=True, color=WHITE, size=18, name="Arial")
    cell.fill = PatternFill("solid", start_color=DARK_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 40

    r += 1
    cell = ws.cell(row=r, column=1, value="UpTopFlow LLC — Mission Critical Facility Documentation")
    cell.font = Font(bold=False, color=WHITE, size=11, name="Arial")
    cell.fill = PatternFill("solid", start_color=TEAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 20

    r += 2
    section_header(ws, r, "DOCUMENT IDENTIFICATION")
    r += 1
    fields = [
        ("MOP Title", meta.get("title", "")),
        ("MOP Number", meta.get("number", "")),
        ("System", meta.get("system", "")),
        ("Facility", meta.get("facility", "")),
        ("Revision Date", meta.get("date", str(date.today()))),
        ("Written By", "UpTopFlow LLC"),
    ]
    for label, value in fields:
        label_cell(ws, r, 1, label, span=2)
        fill_cell(ws, r, 3, value, span=6, bg=WHITE)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_header(ws, r, "ACTIVITY SUMMARY")
    r += 1
    summary_fields = [
        ("Project / Activity Name", ""),
        ("Planned Date", ""),
        ("Estimated Start Time", ""),
        ("Estimated Completion Time", ""),
        ("Actual Start Time", ""),
        ("Actual Completion Time", ""),
        ("Scope of Work", ""),
        ("Participating Parties", ""),
        ("Key Risks", ""),
        ("Mitigation", ""),
        ("Contingency Plan", ""),
    ]
    for label, value in summary_fields:
        label_cell(ws, r, 1, label, span=2)
        fill_cell(ws, r, 3, value, span=6)
        ws.row_dimensions[r].height = 30 if label in ("Scope of Work", "Key Risks", "Mitigation", "Contingency Plan") else 18
        r += 1

    r += 1
    section_header(ws, r, "APPROVAL AND SIGN-OFF")
    r += 1
    ws.row_dimensions[r].height = 16
    for col, heading in enumerate(["Role", "Name", "Signature", "Date"], start=1):
        label_cell(ws, r, col * 2 - 1, heading, span=2)
    r += 1
    approval_roles = [
        "Prepared By",
        "Technically Reviewed By",
        "Safety Reviewed By",
        "Approved By",
        "Authorized By",
    ]
    for role in approval_roles:
        label_cell(ws, r, 1, role, span=2, bold=False)
        for col in [3, 5, 7]:
            fill_cell(ws, r, col, span=2)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1
    section_header(ws, r, "REQUIRED PERSONNEL")
    r += 1
    ws.row_dimensions[r].height = 16
    for col, heading in enumerate(["Role", "Assigned Person", "Contact Number", "Minimum Qualification"], start=1):
        label_cell(ws, r, col * 2 - 1, heading, span=2)
    r += 1
    personnel_roles = meta.get("personnel_roles", [
        "Team 1 Control Lead",
        "Team 2 Switching",
        "Team 3 Verification",
        "Team 4 Floor Patrol",
        "Vendor Lead",
    ])
    for role in personnel_roles:
        label_cell(ws, r, 1, role, span=2, bold=False)
        for col in [3, 5]:
            fill_cell(ws, r, col, span=2)
        fill_cell(ws, r, 7, "[Verify qualification]", span=2, bg=WHITE)
        ws.row_dimensions[r].height = 20
        r += 1


# ── Sheet 2: Safety & Prerequisites ──────────────────────────────────────────
def build_safety_sheet(ws, warnings, prerequisites, ppe):
    ws.title = "Safety & Prerequisites"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 80

    r = 1
    section_header(ws, r, "SAFETY WARNINGS — READ BEFORE STARTING", num_cols=2)
    r += 1
    cell = ws.cell(row=r, column=1,
                   value="⚠ These warnings must be reviewed and acknowledged by all personnel before any step is executed.")
    cell.font = Font(bold=True, color=RED_WARN, size=10, name="Arial")
    cell.fill = PatternFill("solid", start_color="FDECEA")
    cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 20
    r += 1

    for i, warning in enumerate(warnings, start=1):
        ws.cell(row=r, column=1, value=f"⚠ {i}")
        ws.cell(row=r, column=1).font = Font(bold=True, color=RED_WARN,
                                              size=10, name="Arial")
        ws.cell(row=r, column=1).fill = PatternFill("solid",
                                                     start_color="FDECEA")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center",
                                                        vertical="top")
        warn_cell(ws, r, 2, warning)
        ws.row_dimensions[r].height = max(30, len(warning) // 3)
        r += 1

    r += 1
    section_header(ws, r, "REQUIRED PPE", num_cols=2)
    r += 1
    for line in ppe.splitlines():
        if line.strip():
            label_cell(ws, r, 1, "•", bold=False, span=1, bg=WHITE)
            fill_cell(ws, r, 2, line.strip(), bg=WHITE)
            ws.row_dimensions[r].height = 28
            r += 1

    r += 1
    section_header(ws, r, "PREREQUISITES — CONFIRM ALL BEFORE STARTING", num_cols=2)
    r += 1
    label_cell(ws, r, 1, "✓", bold=True)
    label_cell(ws, r, 2, "Prerequisite", bold=True)
    r += 1
    for line in prerequisites.splitlines():
        line = line.strip().lstrip("0123456789.-) ")
        if line:
            fill_cell(ws, r, 1, "□", bg=YELLOW_FILL)
            fill_cell(ws, r, 2, line, bg=WHITE)
            ws.row_dimensions[r].height = 32
            r += 1


# ── Sheet 3: Procedure Steps ──────────────────────────────────────────────────
def build_procedure_sheet(ws, steps):
    ws.title = "Procedure Steps"
    ws.sheet_view.showGridLines = False

    col_widths = [8, 6, 55, 18, 18, 18]
    col_letters = ["A", "B", "C", "D", "E", "F"]
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    r = 1
    section_header(ws, r, "STEP BY STEP PROCEDURE", num_cols=6)
    r += 1

    # Column headers
    headers = ["Step #", "✓", "Action", "Completed By", "Time", "Notes"]
    for col, h in enumerate(headers, start=1):
        label_cell(ws, r, col, h)
    ws.row_dimensions[r].height = 28
    r += 1

    for step in steps:
        num = step["number"]
        text = step["text"]

        if step.get("is_phase"):
            # Phase header row — spans full width, teal background
            section_header(ws, r, text, num_cols=6)
            ws.row_dimensions[r].height = 24
            r += 1
            continue

        if step["is_ponr"]:
            bg = ORANGE_FILL
            prefix = "⛔ POINT OF NO RETURN — "
            step_num_display = "!"
        elif step["is_verify"]:
            bg = GREEN_FILL
            prefix = "✅ VERIFY — "
            step_num_display = "✓"
        elif step["is_photo"]:
            bg = "E3F2FD"
            prefix = "📷 "
            step_num_display = "📷"
        else:
            bg = WHITE if (num or 0) % 2 == 0 else LIGHT_GRAY
            prefix = ""
            step_num_display = str(num) if num else ""

        # Step number cell
        num_cell = ws.cell(row=r, column=1, value=step_num_display)
        num_cell.font = Font(bold=True, size=10, name="Arial")
        num_cell.alignment = Alignment(horizontal="center", vertical="top")
        num_cell.fill = PatternFill("solid", start_color=bg)
        num_cell.border = thin_border()

        # Checkbox cell
        fill_cell(ws, r, 2, "□", bg=YELLOW_FILL)

        # Action cell
        action_cell = ws.cell(row=r, column=3, value=prefix + text)
        action_cell.font = Font(
            bold=step["is_ponr"] or step["is_verify"],
            size=10,
            name="Arial",
            color=RED_WARN if step["is_ponr"] else "000000"
        )
        action_cell.fill = PatternFill("solid", start_color=bg)
        action_cell.alignment = Alignment(wrap_text=True, vertical="top")
        action_cell.border = thin_border()

        # Fillable columns — Completed By, Time, Notes
        for col in [4, 5, 6]:
            fill_cell(ws, r, col, bg=YELLOW_FILL)

        estimated_height = max(40, len(text) // 2)
        ws.row_dimensions[r].height = estimated_height
        r += 1

    r += 1
    section_header(ws, r, "EXPECTED OUTCOME", num_cols=6)
    r += 1
    fill_cell(ws, r, 1, "", span=6, bg=WHITE)
    ws.row_dimensions[r].height = 40

    r += 1
    section_header(ws, r, "IF SOMETHING GOES WRONG", num_cols=6)
    r += 1
    fill_cell(ws, r, 1, "", span=6, bg="FDECEA")
    ws.row_dimensions[r].height = 40

    r += 1
    section_header(ws, r, "ESCALATION CONTACTS", num_cols=6)
    r += 1
    for contact in ["Primary", "Backup", "Facility Engineer", "Vendor Lead"]:
        label_cell(ws, r, 1, contact, span=2)
        fill_cell(ws, r, 3, span=4)
        ws.row_dimensions[r].height = 28
        r += 1


# ── Sheet 4: Checklists ───────────────────────────────────────────────────────
def build_checklist_sheet(ws):
    ws.title = "Checklists"
    ws.sheet_view.showGridLines = False

    for col, width in zip(range(1, 7), [6, 35, 20, 15, 15, 20]):
        ws.column_dimensions[get_column_letter(col)].width = width

    r = 1

    # LOTO checklist
    section_header(ws, r, "APPENDIX B — LOTO CHECKLIST", num_cols=6)
    r += 1
    label_cell(ws, r, 1, "✓", bold=True)
    for col, h in enumerate(["Isolation Point", "Applied By", "Lock/Tag #",
                              "Time Applied", "Authorized By"], start=2):
        label_cell(ws, r, col, h)
    r += 1
    for _ in range(8):
        fill_cell(ws, r, 1, "□", bg=YELLOW_FILL)
        for col in range(2, 7):
            fill_cell(ws, r, col)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_header(ws, r, "APPENDIX C — VOLTAGE VERIFICATION CHECKLIST", num_cols=6)
    r += 1
    for col, h in enumerate(["Equipment / Location", "Expected Reading",
                              "Actual Reading", "Technician", "Verifier",
                              "Time"], start=1):
        label_cell(ws, r, col, h)
    r += 1
    for _ in range(8):
        for col in range(1, 7):
            fill_cell(ws, r, col)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_header(ws, r, "APPENDIX A — PRE-ACTIVITY INFRASTRUCTURE VERIFICATION", num_cols=6)
    r += 1
    for col, h in enumerate(["✓", "Equipment", "Pre-Activity Condition",
                              "Verified By", "Time", "Notes"], start=1):
        label_cell(ws, r, col, h)
    r += 1
    for _ in range(10):
        fill_cell(ws, r, 1, "□", bg=YELLOW_FILL)
        for col in range(2, 7):
            fill_cell(ws, r, col)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_header(ws, r, "APPENDIX D — POST-ACTIVITY CLOSE-OUT CHECKLIST", num_cols=6)
    r += 1
    closeout_items = [
        "All in-scope equipment restored to normal energized operation",
        "All monitoring systems show normal readings with no active alarms",
        "All tools and test equipment removed from all work areas",
        "All enclosure doors closed and secured",
        "All LOTO devices removed and accounted for",
        "All maintenance records signed and collected",
        "Client notification sent confirming activity complete",
        "Activity log closed with final completion time recorded",
        "All punch list items documented and escalated",
    ]
    label_cell(ws, r, 1, "✓", bold=True)
    label_cell(ws, r, 2, "Item", bold=True, span=3)
    label_cell(ws, r, 5, "Confirmed By", bold=True)
    label_cell(ws, r, 6, "Time", bold=True)
    r += 1
    for item in closeout_items:
        fill_cell(ws, r, 1, "□", bg=YELLOW_FILL)
        fill_cell(ws, r, 2, item, span=3, bg=WHITE)
        fill_cell(ws, r, 5, bg=YELLOW_FILL)
        fill_cell(ws, r, 6, bg=YELLOW_FILL)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1
    label_cell(ws, r, 1, "Final Close-Out Sign-Off:", span=2)
    fill_cell(ws, r, 3, span=4)
    ws.row_dimensions[r].height = 20


# ── Sheet 5: Risk Assessment ──────────────────────────────────────────────────
def build_risk_sheet(ws):
    ws.title = "Risk Assessment"
    ws.sheet_view.showGridLines = False

    for col, width in zip(range(1, 7), [30, 12, 12, 30, 20, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width

    r = 1
    section_header(ws, r, "APPENDIX E — RISK ASSESSMENT", num_cols=6)
    r += 1
    for col, h in enumerate(["Risk", "Likelihood", "Impact",
                              "Mitigation", "Owner", "Status"], start=1):
        label_cell(ws, r, col, h)
    r += 1
    for _ in range(8):
        for col in range(1, 7):
            fill_cell(ws, r, col)
        ws.row_dimensions[r].height = 32
        r += 1


# ── Main converter ────────────────────────────────────────────────────────────
def convert_mop_to_excel(md_path, output_path,
                          facility_name="", personnel_roles=None):
    """
    Reads a generated MOP markdown file and writes a formatted Excel workbook.

    Parameters:
        md_path: Path to the markdown MOP file (e.g. outputs/MOP-POW-001.md)
        output_path: Where to save the Excel file (e.g. outputs/MOP-POW-001.xlsx)
        facility_name: Facility name for the cover page
        personnel_roles: Optional list of role names for the personnel table
    """

    with open(md_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    sections = parse_mop_markdown(md_text)

    # Pull metadata from first lines
    title_match = re.search(r"MOP TITLE:\s*(.+)", md_text)
    number_match = re.search(r"MOP NUMBER:\s*(.+)", md_text)
    system_match = re.search(r"^SYSTEM:\s*(.+)", md_text, re.MULTILINE)
    date_match = re.search(r"REVISION DATE:\s*(.+)", md_text)

    meta = {
        "title": title_match.group(1).strip() if title_match else "",
        "number": number_match.group(1).strip() if number_match else "",
        "system": system_match.group(1).strip() if system_match else "",
        "facility": facility_name,
        "date": date_match.group(1).strip() if date_match else str(date.today()),
        "personnel_roles": personnel_roles or [
            "Team 1 Control Lead",
            "Team 2 Switching",
            "Team 3 Verification",
            "Team 4 Floor Patrol",
            "Vendor Lead",
        ],
    }

    steps_text = sections.get("STEP BY STEP PROCEDURE", "")
    steps = extract_steps(steps_text)

    warnings_text = sections.get("SAFETY WARNINGS", "")
    warnings = extract_warnings(warnings_text)
    if not warnings:
        warnings = ["Review all safety warnings with the full team before execution."]

    prerequisites = sections.get("PREREQUISITES", "No prerequisites extracted — verify with engineer.")
    ppe = sections.get("REQUIRED PPE", "Arc-flash rated PPE required — verify rating with facility arc flash study.")

    wb = Workbook()
    wb.remove(wb.active)

    ws_cover = wb.create_sheet("Cover Page")
    ws_safety = wb.create_sheet("Safety & Prerequisites")
    ws_steps = wb.create_sheet("Procedure Steps")
    ws_checks = wb.create_sheet("Checklists")
    ws_risk = wb.create_sheet("Risk Assessment")

    build_cover_sheet(ws_cover, meta)
    build_safety_sheet(ws_safety, warnings, prerequisites, ppe)
    build_procedure_sheet(ws_steps, steps)
    build_checklist_sheet(ws_checks)
    build_risk_sheet(ws_risk)

    wb.save(output_path)
    print(f"Excel MOP saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    convert_mop_to_excel(
        md_path="outputs/MOP-POW-001.md",
        output_path="outputs/MOP-POW-001.xlsx",
        facility_name="Test Data Center",
        personnel_roles=[
            "Team 1 Control Lead",
            "Team 2 Switching",
            "Team 3 Verification",
            "Team 4 Floor Patrol",
            "Vendor Lead",
        ]
    )